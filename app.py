from flask import Flask, render_template, send_file
from flask_mysqldb import MySQL
from flask import request, redirect, url_for
from io import BytesIO
from openpyxl import Workbook
from wp import hitung_wp
from collections import Counter
from collections import defaultdict
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet
app = Flask(__name__)



app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''   
app.config['MYSQL_DB'] = 'spk_wp'

mysql = MySQL(app)

# @app.route('/')
# def index():
#     cur = mysql.connection.cursor()
#     cur.execute("SELECT * FROM kriteria")
#     data = cur.fetchall()
#     cur.close()
#     return render_template('index.html', data=data)

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/dashboard')
def dashboard():
    cur = mysql.connection.cursor()

    cur.execute("SELECT * FROM alternatif")
    alternatif_db = cur.fetchall()

    cur.execute("SELECT * FROM kriteria")
    kriteria_db = cur.fetchall()

    cur.execute("SELECT alternatif_id, kriteria_id, nilai FROM nilai")
    nilai_db = cur.fetchall()


    cur.execute("SELECT COUNT(*) FROM alternatif")
    total_alternatif = cur.fetchone()[0]

    cur.execute("SELECT nama_alternatif FROM alternatif")
    data_laptop = cur.fetchall()

    cur.execute("SELECT nama_kriteria, bobot FROM kriteria")
    data_kriteria = cur.fetchall()

    cur.close()


    merk_list = [
    row[0].split()[0].strip().capitalize()
    for row in data_laptop
    ]


    merk_counter = Counter(merk_list)
    labels_merk = list(merk_counter.keys())
    jumlah_merk = list(merk_counter.values())

    
    kriteria_tertinggi_tuple = max(data_kriteria, key=lambda x: float(x[1]))

    kriteria_tertinggi = {
        'nama': kriteria_tertinggi_tuple[0],
        'bobot': float(kriteria_tertinggi_tuple[1])
}

   
    nilai = {(a, k): float(v) for a, k, v in nilai_db}

    alternatif = [
        {'id': a[0], 'nama': a[1]}
        for a in alternatif_db
    ]

    kriteria = [
        {'id': k[0], 'nama': k[1], 'bobot': float(k[2]), 'tipe': k[3]}
        for k in kriteria_db
    ]

    ranking = hitung_wp(alternatif, kriteria, nilai)
    merk_data = defaultdict(list)

    for r in ranking:
        merk = r['nama'].split()[0].capitalize()
        merk_data[merk].append(r['preferensi'])

    # hitung rata-rata skor WP per merk
    merk_ranking = []

    for merk, skor_list in merk_data.items():
        rata_rata = sum(skor_list) / len(skor_list)
        merk_ranking.append({
            'merk': merk,
            'skor': round(rata_rata, 4),
            'jumlah': len(skor_list)
        })

    # urutkan merk terbaik
    merk_ranking = sorted(merk_ranking, key=lambda x: x['skor'], reverse=True)
    labels_merk = labels_merk or []
    jumlah_merk = jumlah_merk or []

    
    return render_template(
        'dashboard.html',
        ranking=ranking,
        total_alternatif=total_alternatif,
        labels_merk=labels_merk,
        jumlah_merk=jumlah_merk,
        kriteria_tertinggi=kriteria_tertinggi,
        merk_ranking=merk_ranking
    )



@app.route('/kriteria', methods=['GET', 'POST'])
def kriteria():
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        nama = request.form['nama']
        bobot = request.form['bobot']
        tipe = request.form['tipe']

        cur.execute(
            "INSERT INTO kriteria (nama_kriteria, bobot, tipe) VALUES (%s,%s,%s)",
            (nama, bobot, tipe)
        )
        mysql.connection.commit()

    cur.execute("SELECT * FROM kriteria")
    rows = cur.fetchall()

    kriteria = []
    for k in rows:
        kriteria.append({
            'id': k[0],
            'nama': k[1],
            'bobot': k[2],
            'tipe': k[3]
        })

    return render_template('kriteria.html', kriteria=kriteria)



@app.route('/alternatif', methods=['GET', 'POST'])
def alternatif():
    cur = mysql.connection.cursor()

    if request.method == 'POST':
        nama = request.form['nama']
        cur.execute(
            "INSERT INTO alternatif (nama_alternatif) VALUES (%s)",
            (nama,)
        )
        mysql.connection.commit()
        cur.close()

        
        return redirect(url_for('nilai'))

    
    cur.execute("SELECT * FROM alternatif")
    data = cur.fetchall()
    cur.close()

    return render_template('alternatif.html', alternatif=data)


@app.route('/hasil')
def hasil():
    cur = mysql.connection.cursor()

    
    cur.execute("SELECT * FROM alternatif")
    alternatif_db = cur.fetchall()

    
    cur.execute("SELECT id, nama_kriteria, bobot, tipe FROM kriteria")
    kriteria_db = cur.fetchall()

    
    cur.execute("""
        SELECT alternatif_id, kriteria_id, nilai
        FROM nilai
    """)
    nilai_db = cur.fetchall()
    cur.close()

    # Mapping alternatif
    alternatif = [
        {'id': a[0], 'nama': a[1]}
        for a in alternatif_db
    ]

    # Mapping kriteria
    kriteria = load_kriteria()
    print(kriteria)

    
    nilai = {}
    for alt_id, kri_id, nilai_val in nilai_db:
        nilai[(alt_id, kri_id)] = nilai_val

    # print("NILAI DB:", nilai_db)

    # Hitung WP
    ranking = hitung_wp(alternatif, kriteria, nilai)



    return render_template('hasil.html', ranking=ranking)
    
def load_kriteria():
    cur = mysql.connection.cursor()
    cur.execute("SELECT id, nama_kriteria, bobot, tipe FROM kriteria")
    data = cur.fetchall()
    cur.close()

    return [
        {'id': k[0], 'nama_kriteria': k[1], 'bobot': k[2], 'tipe': k[3]}
        for k in data
    ]



@app.route('/nilai', methods=['GET','POST'])
def nilai():
    cur = mysql.connection.cursor()

    cur.execute("SELECT * FROM alternatif")
    alternatif = cur.fetchall()

    cur.execute("SELECT * FROM kriteria")
    kriteria = cur.fetchall()

    if request.method == 'POST':
        alt_id = request.form['alternatif']

        for k in kriteria:
            nilai_input = request.form.get(f'nilai_{k[0]}')

            if nilai_input:
                cur.execute("""
                    INSERT INTO nilai (alternatif_id, kriteria_id, nilai)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE nilai = %s
                """, (alt_id, k[0], nilai_input, nilai_input))

        mysql.connection.commit()
        cur.close()
        return redirect(url_for('dashboard'))

    cur.close()
    return render_template('nilai.html', alternatif=alternatif, kriteria=kriteria)

# Edit & Update / Delete

@app.route('/alternatif/edit/<int:id>')
def edit_alternatif(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM alternatif WHERE id = %s", (id,))
    data = cur.fetchone()
    cur.close()

    if data is None:
        return redirect(url_for('alternatif'))

    return render_template('alternatif_edit.html', alternatif=data)

@app.route('/alternatif/update/<int:id>', methods=['POST'])
def update_alternatif(id):
    nama = request.form['nama']

    cur = mysql.connection.cursor()
    cur.execute(
        "UPDATE alternatif SET nama_alternatif = %s WHERE id = %s",
        (nama, id)
    )
    mysql.connection.commit()
    cur.close()

    return redirect(url_for('alternatif'))

@app.route('/kriteria/edit/<int:id>')
def edit_kriteria(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM kriteria WHERE id = %s", (id,))
    data = cur.fetchone()
    cur.close()

    if data is None:
        return redirect(url_for('kriteria'))

    return render_template('kriteria_edit.html', kriteria=data)

@app.route('/kriteria/update/<int:id>', methods=['POST'])
def update_kriteria(id):
    nama = request.form['nama']
    bobot = request.form['bobot']
    tipe = request.form['tipe']

    cur = mysql.connection.cursor()
    cur.execute("""
        UPDATE kriteria
        SET nama_kriteria = %s, bobot = %s, tipe = %s
        WHERE id = %s
    """, (nama, bobot, tipe, id))
    mysql.connection.commit()
    cur.close()

    return redirect(url_for('kriteria'))

@app.route('/kriteria/delete/<int:id>', methods=['POST'])
def delete_kriteria(id):
    cur = mysql.connection.cursor()

    # hapus nilai dulu (FK)
    cur.execute("DELETE FROM nilai WHERE kriteria_id = %s", (id,))
    cur.execute("DELETE FROM kriteria WHERE id = %s", (id,))

    mysql.connection.commit()
    cur.close()

    return redirect(url_for('kriteria'))

@app.route('/alternatif/delete/<int:id>', methods=['POST'])
def delete_alternatif(id):
    cur = mysql.connection.cursor()

    # 1. Hapus nilai milik alternatif ini dulu
    cur.execute("DELETE FROM nilai WHERE alternatif_id = %s", (id,))

    # 2. Baru hapus alternatif
    cur.execute("DELETE FROM alternatif WHERE id = %s", (id,))

    mysql.connection.commit()
    cur.close()

    return redirect(url_for('alternatif'))

@app.route('/export/pdf')
def export_pdf():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM alternatif")
    alternatif_db = cur.fetchall()
    cur.execute("SELECT id, nama_kriteria AS nama, bobot, tipe FROM kriteria")
    kriteria_db = cur.fetchall()
    cur.execute("SELECT alternatif_id, kriteria_id, nilai FROM nilai")
    nilai_db = cur.fetchall()
    cur.close()

    alternatif = [{'id': a[0], 'nama': a[1]} for a in alternatif_db]
    kriteria = load_kriteria()
    nilai = {(a, k): n for a, k, n in nilai_db}

    ranking = hitung_wp(alternatif, kriteria, nilai)

    file_path = '/mnt/data/hasil_wp.pdf'
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Hasil Perhitungan Weighted Product", styles['Title']))
    elements.append(Paragraph("<br/>", styles['Normal']))

    # Header tabel
    header = ['Alternatif', 'Skor WP', 'Ranking']
    for k in kriteria:
        header.append(k['nama'])

    table_data = [header]

    for r in ranking:
        row = [
            r['nama'],
            round(r['nilai'], 5),
            r['rank']
        ]
        for k in kriteria:
            row.append(r['detail'].get(k['nama'], '-'))
        table_data.append(row)

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (1,1), (-1,-1), 'CENTER')
    ]))

    elements.append(table)
    doc.build(elements)

    return send_file(
        file_path,
        as_attachment=True,
        download_name='hasil_wp.pdf'
    )

@app.route('/export/excel')
def export_excel():
    cur = mysql.connection.cursor()

    cur.execute("SELECT * FROM alternatif")
    alternatif_db = cur.fetchall()

    cur.execute("SELECT id, nama_kriteria AS nama, bobot, tipe FROM kriteria")
    kriteria_db = cur.fetchall()

    cur.execute("SELECT alternatif_id, kriteria_id, nilai FROM nilai")
    nilai_db = cur.fetchall()
    cur.close()

    alternatif = [{'id': a[0], 'nama': a[1]} for a in alternatif_db]
    kriteria = load_kriteria()
    nilai = {(a, k): n for a, k, n in nilai_db}

    ranking = hitung_wp(alternatif, kriteria, nilai)


    data = []

    for r in ranking:
        row = {
            "Alternatif": r["nama"],
            "Skor WP": round(r["nilai"], 5),
            "Ranking": r["rank"],
        }

        for nama_kriteria, skor in r["detail"].items():
            nama_bersih = normalize_kriteria(nama_kriteria)

            keterangan = KRITERIA_INDEX.get(nama_bersih, {}).get(
                skor, f"Skor {skor}"
            )

            row[nama_bersih] = keterangan


        data.append(row)


# ----- SUSUN DATA EXCEL -----
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil WP"

    # Header otomatis
    headers = list(data[0].keys())
    ws.append(headers)

    # Isi data
    for row in data:
        ws.append([row[h] for h in headers])

    # AUTO WIDTH
    from openpyxl.utils import get_column_letter

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2


    # START_COL = 13   # Kolom M
    # START_ROW = 1

    # ws.cell(row=START_ROW, column=START_COL, value="Kriteria")
    # ws.cell(row=START_ROW, column=START_COL + 1, value="Nilai")
    # ws.cell(row=START_ROW, column=START_COL + 2, value="Keterangan")

    # row_idx = START_ROW + 1

    # for nama_kriteria, mapping in KRITERIA_INDEX.items():
    #     for nilai, ket in mapping.items():
    #         ws.cell(row=row_idx, column=START_COL, value=nama_kriteria)
    #         ws.cell(row=row_idx, column=START_COL + 1, value=nilai)
    #         ws.cell(row=row_idx, column=START_COL + 2, value=ket)
    #         row_idx += 1


    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='hasil_wp.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)

def load_kriteria():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT
            id,
            nama_kriteria AS nama,
            bobot,
            tipe
        FROM kriteria
    """)
    data = cur.fetchall()
    cur.close()

    return [
        {
            'id': k[0],
            'nama': k[1],
            'bobot': float(k[2]),
            'tipe': k[3]
        }
        for k in data
    ]

KRITERIA_INDEX = {
    "Harga": {
        5: "≤ 6 Juta",
        4: "6 – 7.5 Juta",
        3: "7.5 – 9 Juta",
        2: "9 – 11 Juta",
        1: "> 11 Juta",
    },
    "RAM": {
        1: "4 GB",
        2: "8 GB",
        3: "16 GB",
        4: "32 GB",
    },
    "Storage": {
        1: "HDD",
        2: "SSD 128 GB",
        3: "SSD 256 GB",
        4: "SSD 512 GB",
        5: "SSD ≥ 1 TB",
    },
    "Processor": {
        1: "Celeron / Pentium / Athlon",
        2: "i3 Gen 8–9 / Ryzen 3 3000",
        3: "i3 Gen 10+ / Ryzen 3 5000 / i5 Gen 8–9",
        4: "i5 Gen 10+ / i7 Gen 10–11 / Apple M1/M2",
        5: "i7 Gen 12+ / i9 / Ryzen 9 / Apple M2 Max",
    },
    "Berat": {
        5: "≤ 1.5 kg (Sangat ringan)",
        4: "1.6 – 2.0 kg (Ringan)",
        3: "2.1 – 2.5 kg (Sedang)",
        2: "2.6 – 3.0 kg (Berat)",
        1: "> 3.0 kg (Sangat berat)",
    },
}

def normalize_kriteria(nama):
    return (
        nama.replace("[", "")
            .replace("]", "")
            .replace(" Besar", "")
            .strip()
    )



if __name__ == '__main__':
    app.run(debug=True)
